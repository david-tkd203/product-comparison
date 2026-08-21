import os
import json
import re
import urllib.request
import urllib.error
from datetime import datetime
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import wraps
from io import BytesIO

import mysql.connector

from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFProtect

from flask import Flask, render_template, request, redirect, url_for, flash, make_response, jsonify, session
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd

app = Flask(__name__)

app.url_map.strict_slashes = False

csrf = CSRFProtect(app)
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["1000 per day", "200 per hour"],
    storage_uri="memory://"
)

app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'cambiar-en-produccion')

DB_CONFIG = {
    'host': os.environ.get('MYSQL_HOST', 'db'),
    'user': 'perfumes',
    'password': 'perfumes',
    'database': 'perfumes',
    'charset': 'utf8mb4',
    'autocommit': True,
}

HEADER_ROW = 6


def get_db():
    return mysql.connector.connect(**DB_CONFIG)


def _add_column_if_missing(cur, table, column, col_type):
    """Add a column if it doesn't exist in the table."""
    try:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
    except mysql.connector.Error:
        pass  # column already exists


def init_db():
    db = get_db()
    cur = db.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS products (
            sku VARCHAR(50) PRIMARY KEY,
            nombre TEXT,
            linea VARCHAR(100),
            ean VARCHAR(50),
            genero VARCHAR(50),
            formato VARCHAR(50)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''')
    _add_column_if_missing(cur, 'products', 'stock', 'INT DEFAULT 0')
    _add_column_if_missing(cur, 'products', 'is_active', 'TINYINT(1) DEFAULT 0')
    _add_column_if_missing(cur, 'products', 'aroma_family', 'VARCHAR(50) DEFAULT NULL')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS prices (
            id INT AUTO_INCREMENT PRIMARY KEY,
            sku VARCHAR(50) NOT NULL,
            import_date DATE NOT NULL,
            precio INT NOT NULL,
            FOREIGN KEY (sku) REFERENCES products(sku)
        ) ENGINE=InnoDB
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS imports (
            id INT AUTO_INCREMENT PRIMARY KEY,
            filename TEXT,
            import_date DATE UNIQUE,
            product_count INT
        ) ENGINE=InnoDB
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS cosmetic_products (
            sku VARCHAR(50) PRIMARY KEY,
            nombre TEXT,
            precio_retail INT,
            precio_ref INT,
            imagen TEXT,
            url TEXT,
            body_html LONGTEXT,
            aromas TEXT,
            last_synced TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''')
    _add_column_if_missing(cur, 'cosmetic_products', 'body_html', 'LONGTEXT')
    _add_column_if_missing(cur, 'cosmetic_products', 'aromas', 'TEXT')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS silk_products (
            sku VARCHAR(50) PRIMARY KEY,
            nombre TEXT,
            precio_retail INT,
            precio_ref INT,
            imagen TEXT,
            url TEXT,
            body_html LONGTEXT,
            last_synced TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''')
    _add_column_if_missing(cur, 'silk_products', 'body_html', 'LONGTEXT')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS silk_matches (
            sku_wholesale VARCHAR(50) PRIMARY KEY,
            sku_silk VARCHAR(50),
            confidence FLOAT
        ) ENGINE=InnoDB
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS multimarca_products (
            sku VARCHAR(50) PRIMARY KEY,
            nombre TEXT,
            precio_retail INT,
            precio_ref INT,
            imagen TEXT,
            url TEXT,
            body_html LONGTEXT,
            last_synced TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''')
    _add_column_if_missing(cur, 'multimarca_products', 'body_html', 'LONGTEXT')
    _add_column_if_missing(cur, 'multimarca_products', 'aromas', 'TEXT')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS multimarca_matches (
            sku_wholesale VARCHAR(50) PRIMARY KEY,
            sku_mm VARCHAR(50),
            confidence FLOAT
        ) ENGINE=InnoDB
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS catalog_links (
            id INT AUTO_INCREMENT PRIMARY KEY,
            token VARCHAR(32) UNIQUE NOT NULL,
            margen_pct DECIMAL(5,2) NOT NULL DEFAULT 30.00,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP NULL,
            active TINYINT(1) DEFAULT 1
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INT AUTO_INCREMENT PRIMARY KEY,
            link_token VARCHAR(32) NOT NULL,
            nombre VARCHAR(200) NOT NULL,
            telefono VARCHAR(50) NOT NULL,
            items_json LONGTEXT NOT NULL,
            total INT NOT NULL DEFAULT 0,
            status VARCHAR(20) DEFAULT 'nuevo',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_token (link_token),
            INDEX idx_status (status)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''')

    # ponytail: dedup + UNIQUE on (sku, import_date). Without it, re-uploads
    # create duplicate price rows and every JOIN multiplies rows Nx.
    try:
        cur.execute('''SELECT COUNT(*) FROM (
            SELECT sku, import_date, COUNT(*) as cnt FROM prices
            GROUP BY sku, import_date HAVING cnt > 1
        ) dupes''')
        dupes = cur.fetchone()[0]
    except Exception:
        dupes = 0

    if dupes > 0:
        autocommit = db.autocommit
        db.autocommit = False
        try:
            cur.execute('''CREATE TABLE prices_dedup LIKE prices''')
            cur.execute('''ALTER TABLE prices_dedup ADD UNIQUE INDEX unique_sku_date (sku, import_date)''')
            cur.execute('''INSERT IGNORE INTO prices_dedup (sku, import_date, precio)
                           SELECT sku, import_date, MAX(precio) FROM prices
                           GROUP BY sku, import_date''')
            cur.execute('''RENAME TABLE prices TO prices_old, prices_dedup TO prices''')
            cur.execute('''DROP TABLE prices_old''')
            db.commit()
        except Exception:
            db.rollback()
            raise
        finally:
            db.autocommit = autocommit
    else:
        try:
            cur.execute('''ALTER TABLE prices ADD UNIQUE INDEX unique_sku_date (sku, import_date)''')
        except Exception:
            pass  # index already exists

    # Users table for admin login
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''')

    # Create or update default admin password from environment
    admin_pass = os.environ.get('ADMIN_PASSWORD', 'admin123')
    cur.execute('SELECT COUNT(*) as cnt FROM users')
    if cur.fetchone()[0] == 0:
        cur.execute('INSERT INTO users (username, password_hash) VALUES (%s, %s)',
                    ['admin', generate_password_hash(admin_pass)])
    else:
        # Ensure admin password always matches current ADMIN_PASSWORD env var
        cur.execute('UPDATE users SET password_hash = %s WHERE username = %s',
                    [generate_password_hash(admin_pass), 'admin'])

    # Global key/value settings (e.g. store margin)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            setting_key VARCHAR(50) PRIMARY KEY,
            setting_value TEXT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''')

    cur.close()
    db.close()


DEFAULT_SETTINGS = {'margen_pct': '30'}


def get_setting(db, key, default=None):
    row = _query(db, 'SELECT setting_value FROM settings WHERE setting_key = %s', [key]).fetchone()
    if row and row['setting_value'] is not None:
        return row['setting_value']
    return default if default is not None else DEFAULT_SETTINGS.get(key)


def set_setting(db, key, value):
    _query(db, '''INSERT INTO settings (setting_key, setting_value) VALUES (%s, %s)
                  ON DUPLICATE KEY UPDATE setting_value = VALUES(setting_value)''',
           [key, str(value)])


# Gender values in the wholesale XLSX are messy (HOMBRE, Masculino, Women...).
# Normalize to the three canonical buckets used by the storefront.
GENERO_FEMALE = ('mujer', 'femen', 'dama', 'female', 'women', 'señora', 'senora')
GENERO_MALE = ('hombre', 'mascul', 'caball', 'male', 'men', 'señor', 'senor', 'él', 'el hombre')


def _norm_genero(g):
    if not g:
        return 'Unisex'
    g = str(g).strip().lower()
    if any(k in g for k in GENERO_FEMALE):
        return 'Mujer'
    if any(k in g for k in GENERO_MALE):
        return 'Hombre'
    return 'Unisex'


def _genero_sql_conds(genero_norm):
    """Parametrized LIKE conditions matching raw genero values to a normalized bucket."""
    if genero_norm == 'Mujer':
        pats = [f'%{k}%' for k in GENERO_FEMALE]
    elif genero_norm == 'Hombre':
        pats = [f'%{k}%' for k in GENERO_MALE]
    else:
        return [], []
    conds = ['LOWER(p.genero) LIKE %s'] * len(pats)
    return conds, pats


def require_login(f):
    """Decorator to protect admin routes."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Iniciá sesión para continuar', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def parse_excel(filepath):
    df = pd.read_excel(filepath, engine='openpyxl', header=None)

    # Detect the header row — look for the row that has the most expected keywords
    TARGET_KEYWORDS = ['sku', 'nombre', 'precio', 'linea']
    best_row = HEADER_ROW
    best_score = 0
    for r in range(min(12, len(df))):
        row_vals = [str(v).strip().lower().replace('\ufffd', '') for v in df.iloc[r].tolist() if isinstance(v, str) or not pd.isna(v)]
        row_text = ' '.join(row_vals)
        score = sum(1 for kw in TARGET_KEYWORDS if kw in row_text)
        if score > best_score:
            best_score = score
            best_row = r

    headers = [str(h).strip() for h in df.iloc[best_row].tolist()]
    data = df.iloc[best_row + 1:].copy()
    data.columns = headers

    # Normalize: NFC normalize, lowercase, strip accents, replace encoding artifacts
    import unicodedata
    def _norm(s):
        s = str(s).strip()
        s = unicodedata.normalize('NFKD', s)          # decompose accents
        s = s.lower()
        s = ''.join(c for c in s if not unicodedata.combining(c))  # strip combining marks
        s = s.replace('\ufffd', '')                    # strip replacement chars
        return re.sub(r'[^a-z0-9 ]', '', s)            # keep only alphanumeric
    data.columns = [_norm(c) for c in data.columns]

    # Robust column mapping: match normalized names to target columns
    COLUMN_ALIASES = {
        'linea':   ['linea', 'marca', 'brand', 'marcas', 'line'],
        'nombre':  ['nombre', 'name', 'producto', 'product', 'descripcion', 'description', 'desc', 'titulo', 'title', 'item'],
        'sku':     ['sku', 'codigo', 'code', 'cod', 'id', 'referencia', 'ref', 'reference'],
        'ean':     ['ean', 'upc', 'barcode', 'codigo de barras', 'codbarra', 'cod barras', 'gtin', 'codigobarras'],
        'genero':  ['genero', 'gender', 'sexo', 'gener', 'gnero'],
        'formato': ['formato', 'format', 'presentacion', 'tamano', 'tamanio', 'tam', 'tipo', 'talla', 'size', 'volumen', 'ml'],
        'precio':  ['precio', 'price', 'precio mayorista', 'costo', 'cost', 'valor', 'precio unitario',
                     'p mayorista', 'precio_mayorista', 'p unitario', 'p neto', 'neto', 'importe',
                     'total', 'valorneto', 'precioneto'],
    }

    rename_map = {}
    found = set()
    for target, aliases in COLUMN_ALIASES.items():
        for col in data.columns:
            if col in aliases or any(a in col for a in aliases):
                rename_map[col] = target
                found.add(target)
                break

    missing = set(COLUMN_ALIASES.keys()) - found
    if missing:
        raise ValueError(
            f'No se encontraron las columnas: {", ".join(missing)}. '
            f'Columnas detectadas: {", ".join(data.columns.tolist())}. '
            f'El archivo debe tener columnas como: SKU, Nombre, Precio, Linea.'
        )

    data = data.rename(columns=rename_map)
    data = data[['linea', 'nombre', 'sku', 'ean', 'genero', 'formato', 'precio']]
    data = data.dropna(subset=['sku'])
    data['sku'] = data['sku'].astype(str).str.strip()
    data['precio'] = pd.to_numeric(data['precio'], errors='coerce').fillna(0).astype(int)
    for col in ['nombre', 'linea', 'ean', 'genero', 'formato']:
        data[col] = data[col].astype(str).str.strip()
    return data


@app.context_processor
def inject_tunnel_url():
    return dict(tunnel_url='')


def _query(db, sql, params=None):
    cur = db.cursor(dictionary=True)
    cur.execute(sql, params or ())
    return cur


def _paginate(sql, params, page, per_page):
    """Return paginated results plus total count."""
    page = max(1, page)
    offset = (page - 1) * per_page
    # Count total
    count_sql = f'SELECT COUNT(*) AS cnt FROM ({sql}) AS t'
    # Remove LIMIT clause from count query if present in the raw sql
    # The caller should pass the base SQL without LIMIT.
    # We append LIMIT/OFFSET to data query separately.
    data_sql = f'{sql} LIMIT %s OFFSET %s'
    data_params = list(params) + [per_page, offset]
    return data_sql, data_params, page, per_page, offset



@app.route('/admin')
@app.route('/admin/')
@require_login
def admin_redirect():
    return redirect(url_for('admin_productos'))

@app.route('/admin/productos')
@require_login
def admin_productos():
    query = request.args.get('q', '').strip()
    linea = request.args.get('linea', '').strip()
    genero = request.args.get('genero', '').strip()
    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = 50

    db = get_db()
    lineas = [r['linea'] for r in _query(db, 'SELECT DISTINCT linea FROM products ORDER BY linea')]
    generos = [r['genero'] for r in _query(db, 'SELECT DISTINCT genero FROM products ORDER BY genero')]
    latest = _query(db, 'SELECT MAX(import_date) as max_date FROM imports').fetchone()['max_date']

    sql = '''SELECT p.sku, p.nombre, p.linea, p.ean, p.genero, p.formato, pr.precio
             FROM products p
             JOIN prices pr ON p.sku = pr.sku'''
    params = []
    conditions = []
    if latest:
        conditions.append('pr.import_date = %s')
        params.append(latest)
    if query:
        conditions.append('(p.nombre LIKE %s OR p.sku LIKE %s OR p.linea LIKE %s)')
        like = f'%{query}%'
        params.extend([like, like, like])
    if linea:
        conditions.append('p.linea = %s')
        params.append(linea)
    if genero:
        conditions.append('p.genero = %s')
        params.append(genero)
    if conditions:
        sql += ' WHERE ' + ' AND '.join(conditions)
    sql += ' ORDER BY p.linea, p.nombre'

    # Count total
    total = _query(db, f'SELECT COUNT(*) AS cnt FROM ({sql}) AS t', params).fetchone()['cnt']
    offset = (page - 1) * per_page
    products = _query(db, f'{sql} LIMIT %s OFFSET %s', params + [per_page, offset]).fetchall()
    total_pages = (total + per_page - 1) // per_page if total else 1
    db.close()
    return render_template('admin_productos.html', products=products, query=query,
                          linea=linea, genero=genero, lineas=lineas,
                          generos=generos, latest=latest, total=total,
                          page=page, per_page=per_page, total_pages=total_pages)


@app.route('/upload', methods=['GET', 'POST'])
@require_login
def upload():
    if request.method == 'POST':
        file = request.files.get('file')
        import_date_str = request.form.get('import_date', '').strip()
        if not file or not file.filename.endswith('.xlsx'):
            flash('Subí un archivo .xlsx válido', 'error')
            return render_template('upload.html')
        if import_date_str:
            try:
                import_date = datetime.strptime(import_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Fecha inválida. Usá YYYY-MM-DD', 'error')
                return render_template('upload.html')
        else:
            import_date = datetime.today().date()

        filepath = os.path.join('uploads', file.filename)
        os.makedirs('uploads', exist_ok=True)
        file.save(filepath)

        try:
            data = parse_excel(filepath)
        except Exception as e:
            flash(f'Error al leer el Excel: {e}', 'error')
            return render_template('upload.html')

        db = get_db()
        try:
            existing = _query(db, 'SELECT id FROM imports WHERE import_date = %s', [import_date]).fetchone()
            if existing:
                flash(f'Ya existe una importación con fecha {import_date}.', 'error')
                db.close()
                return render_template('upload.html')

            count = 0
            for _, row in data.iterrows():
                fams = _classify_family({'salida': [str(row['nombre'])]})
                aroma = fams[0] if fams else None
                _query(db, '''INSERT INTO products (sku, nombre, linea, ean, genero, formato, aroma_family)
                              VALUES (%s, %s, %s, %s, %s, %s, %s)
                              ON DUPLICATE KEY UPDATE nombre=VALUES(nombre), linea=VALUES(linea),
                              ean=VALUES(ean), genero=VALUES(genero), formato=VALUES(formato),
                              aroma_family=COALESCE(aroma_family, VALUES(aroma_family))''',
                       [row['sku'], row['nombre'], row['linea'], row['ean'],
                        _norm_genero(row['genero']), row['formato'], aroma])
                _query(db, 'INSERT INTO prices (sku, import_date, precio) VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE precio = VALUES(precio)',
                       [row['sku'], import_date, row['precio']])
                count += 1

            _query(db, 'INSERT INTO imports (filename, import_date, product_count) VALUES (%s, %s, %s)',
                   [file.filename, import_date, count])
            flash(f'Importados {count} productos con fecha {import_date}', 'success')
        except Exception as e:
            flash(f'Error: {e}', 'error')
        finally:
            db.close()
        return redirect(url_for('admin_productos'))
    return render_template('upload.html')


@app.route('/compare')
@require_login
def compare():
    db = get_db()
    imports = _query(db, 'SELECT import_date, filename, product_count FROM imports ORDER BY import_date DESC').fetchall()
    a_date = request.args.get('a', '')
    b_date = request.args.get('b', '')
    diffs = []
    summary = None
    if a_date and b_date:
        diffs = _query(db, '''
            SELECT p.sku, p.nombre, p.linea, p.genero, p.formato,
                   pa.precio as precio_a, pb.precio as precio_b,
                   (pb.precio - pa.precio) as diff,
                   CASE WHEN pa.precio > 0 THEN ROUND((pb.precio - pa.precio) * 100.0 / pa.precio, 1) END as pct
            FROM products p
            JOIN prices pa ON p.sku = pa.sku AND pa.import_date = %s
            JOIN prices pb ON p.sku = pb.sku AND pb.import_date = %s
        ''', [a_date, b_date]).fetchall()

        total = len(diffs)
        subieron = sum(1 for d in diffs if d['diff'] > 0)
        bajaron = sum(1 for d in diffs if d['diff'] < 0)
        iguales = sum(1 for d in diffs if d['diff'] == 0)
        nuevos = _query(db, '''SELECT COUNT(*) as cnt FROM prices pb WHERE pb.import_date = %s
                              AND pb.sku NOT IN (SELECT sku FROM prices WHERE import_date = %s)''',
                        [b_date, a_date]).fetchone()['cnt']
        bajas = _query(db, '''SELECT COUNT(*) as cnt FROM prices pa WHERE pa.import_date = %s
                              AND pa.sku NOT IN (SELECT sku FROM prices WHERE import_date = %s)''',
                       [a_date, b_date]).fetchone()['cnt']
        summary = {
            'total': total, 'subieron': subieron, 'bajaron': bajaron,
            'iguales': iguales, 'nuevos': nuevos, 'bajas': bajas,
            'a_date': a_date, 'b_date': b_date
        }
    db.close()
    return render_template('compare.html', imports=imports, diffs=diffs,
                          summary=summary, a_date=a_date, b_date=b_date)


def _sync_shopify(db, table, base_url):
    added = 0
    error = None
    page = 1
    batch = []
    sql = f'''INSERT INTO {table} (sku, nombre, precio_retail, precio_ref, imagen, url, body_html)
              VALUES (%s, %s, %s, %s, %s, %s, %s)
              ON DUPLICATE KEY UPDATE nombre=VALUES(nombre),
              precio_retail=VALUES(precio_retail), precio_ref=VALUES(precio_ref),
              imagen=VALUES(imagen), url=VALUES(url), body_html=VALUES(body_html),
              last_synced=CURRENT_TIMESTAMP'''

    def _flush():
        nonlocal batch
        if batch:
            cur = db.cursor()
            cur.executemany(sql, batch)
            cur.close()
            batch = []

    while True:
        url = f'{base_url}/products.json?limit=250&page={page}'
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'application/json, text/plain, */*',
                'Accept-Language': 'es-CL,es;q=0.9',
            })
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read())
        except urllib.error.HTTPError as e:
            error = f'HTTP {e.code}: {e.reason}'
            break
        except urllib.error.URLError as e:
            error = f'Error de conexión: {e.reason}'
            break
        except Exception as e:
            error = f'Error inesperado: {e}'
            break
        products = data.get('products', [])
        if not products:
            break
        for p in products:
            for v in p.get('variants', []):
                sku = (v.get('sku') or '').strip()
                if not sku:
                    continue
                precio = int(v.get('price', 0))
                compare_at = int(v.get('compare_at_price') or 0)
                nombre = p.get('title', '')
                url_slug = f'{base_url}/products/{p.get("handle", "")}'
                imagen = (p.get('images') or [{}])[0].get('src', '') if p.get('images') else ''
                body_html = p.get('body_html', '')
                batch.append([sku, nombre, precio, compare_at, imagen, url_slug, body_html])
                added += 1
                if len(batch) >= 100:
                    _flush()
        page += 1
    _flush()
    return (added, error)


# --- Olfactory family classifier ---
AROMA_FAMILIES = {
    'Cítrica': ['limón', 'lima', 'naranja', 'pomelo', 'bergamota', 'mandarina', 'toronja',
                'cítrico', 'citrus', 'cidra', 'yuzu', 'neroli', 'petitgrain', 'verbena',
                'tangerina', 'clementina', 'kumquat', 'limoncello', 'citronela', 'calamansí',
                'naranja sanguina'],
    
    'Floral': ['rosa', 'jazmín', 'violeta', 'lirio', 'flor', 'gardenia', 'magnolia',
               'peonía', 'peonia', 'azahar', 'geranio', 'ylang', 'azucena', 'fresia',
               'narciso', 'jacinto', 'mimosa', 'camelia', 'loto', 'orquídea', 'tuberosa',
               'lavanda', 'madreselva', 'iris', 'clavel', 'crisantemo', 'dalia',
               'floral', 'caléndula', 'margarita', 'heliotropo', 'frangipani', 'nardos',
               'flor de loto', 'flor de almendro', 'flor de cerezo', 'osmanthus', 'pétalo',
               'amapola', 'flor de manzano', 'flor del peral'],
               
    'Amaderada': ['cedro', 'sándalo', 'pino', 'pachulí', 'patchouli', 'madera', 'vetiver',
                  'roble', 'abedul', 'caoba', 'secuoya', 'palo', 'sándal', 'guayaco',
                  'ébano', 'nogal', 'teca', 'cachemira', 'oud', 'agarwood', 'akigalawood',
                  'abeto', 'cálamo', 'ciprés', 'musgo', 'encina', 'haya', 'arce', 'bambú',
                  'maderas', 'woody', 'bois', 'bosque'],
                  
    'Oriental / Especiada': ['vainilla', 'ámbar', 'incienso', 'mirra', 'benjuí', 'canela', 'clavo',
                 'nuez moscada', 'cardamomo', 'pimienta', 'almizcle', 'resina', 'opoponax',
                 'bálsamo', 'tolú', 'estoraque', 'azafrán', 'comino', 'cúrcuma',
                 'especia', 'especiado', 'ambarado', 'oriental', 'almizcl', 'ambreta',
                 'anís', 'regaliz', 'inciens', 'tabaco', 'cumarina', 'haba tonka',
                 'cachemir', 'jengibre', 'cilantro', 'pimentón', 'masala', 'clavo de olor'],
                 
    'Fougère / Aromática': ['fougère', 'fougere', 'helecho', 'musgo de roble', 'salvia',
                'hierba', 'herbal', 'tomillo', 'romero', 'albahaca', 'aromática', 'aromático',
                'artemisa', 'abrótano', 'ajenjo', 'absenta', 'eucalipto', 'menta', 'yerbabuena',
                'laurel', 'hoja', 'té', 'mate', 'orégano', 'estragón', 'angélica', 'hinojo'],
                
    'Chipre / Cuero': ['chipre', 'chypre', 'cuero', 'gamuz', 'gamuza', 'brea', 'alquitrán',
               'ahumado', 'phenol', 'castóreo', 'civet', 'algalia', 'coriáceo',
               'labdanum', 'jara', 'musgo de encina', 'abedul', 'leather', 'cuir', 'humo',
               'tabaco rubio', 'incienso ahumado'],
               
    'Gourmand / Frutal': ['caramelo', 'chocolate', 'café', 'miel', 'azúcar', 'praliné', 'avellana',
                 'cacao', 'dulce', 'goloso', 'gourmand', 'almendra', 'coco', 'leche',
                 'nata', 'crema', 'caramel', 'toffee', 'mazapán', 'turrón', 'chicle',
                 'algodón de azúcar', 'galleta', 'ron', 'whisky', 'licor',
                 'cereza', 'frutilla', 'fresa', 'frambuesa', 'arándano', 'mora',
                 'manzana', 'pera', 'melocotón', 'durazno', 'ciruela', 'cassis',
                 'grosella', 'piña', 'maracuyá', 'mango', 'higo', 'sandía', 'melón',
                 'fruta', 'frutal', 'kiwi', 'papaya', 'guayaba', 'lichi', 'zarzamora',
                 'champaña', 'cognac', 'amaretto'],
                 
    'Acuática / Fresca': ['mar', 'marino', 'agua', 'acuática', 'acuático', 'brisa', 'ozono',
                 'salado', 'sal', 'algas', 'calone', 'cascalone', 'hielo', 'glaciar',
                 'lluvia', 'rocío', 'fresco', 'acuoso', 'loto de agua', 'nenúfar',
                 'bambú de agua']
}


def _classify_family(notas_dict):
    """Classify a perfume's notes into olfactory families."""
    families = set()
    all_notes = []
    for key in ('salida', 'corazon', 'fondo', 'general'):
        all_notes.extend(notas_dict.get(key, []))
    all_text = ' '.join(all_notes).lower()
    for family, keywords in AROMA_FAMILIES.items():
        for kw in keywords:
            if kw in all_text:
                families.add(family)
                break  # one keyword match is enough per family
    return sorted(families)


def _extract_aromas(db, table='cosmetic_products'):
    """Parse body_html to extract fragrance notes. Works for any retail product table."""
    rows = _query(db, f"SELECT sku, body_html, nombre FROM {table} WHERE body_html IS NOT NULL AND body_html != '' AND aromas IS NULL").fetchall()
    extracted = 0
    for r in rows:
        html = r['body_html']
        text = re.sub(r'<[^>]+>', ' ', html)
        text = re.sub(r'\s+', ' ', text).strip()
        
        notas = {'salida': [], 'corazon': [], 'fondo': [], 'general': []}
        
        # 1. Structured notes
        block_start = re.search(r'(?:las?\s+)?(?:contiene\s+)?(?:con\s+)?notas?\s+de\s+salida', text, re.IGNORECASE)
        if block_start:
            block = text[block_start.start():]
            m_salida = re.search(r'notas?\s+de\s+salida\s*(?:son|de|:)?\s*(.+?)(?=\s*(?:las?\s+)?(?:la\s+)?(?:con\s+)?(?:y\s+)?notas?\s+d[ee]l?\s+coraz|\s*(?:las?\s+)?(?:con\s+)?(?:y\s+)?notas?\s+de\s+fondo|\s*$)', block, re.IGNORECASE)
            if m_salida: notas['salida'] = _parse_notas_list(m_salida.group(1))
            m_corazon = re.search(r'notas?\s+d[ee]l?\s+coraz[oó]n\s*(?:son|de|:|es)?\s*(.+?)(?=\s*(?:las?\s+)?(?:con\s+)?(?:y\s+)?notas?\s+de\s+fondo|\s*$)', block, re.IGNORECASE)
            if m_corazon: notas['corazon'] = _parse_notas_list(m_corazon.group(1))
            m_fondo = re.search(r'notas?\s+de\s+fondo\s*(?:son|de|:)?\s*(.+?)(?=\s*(?:<|\.\s*[A-ZÁÉÍÓÚ]|\s*\n\s*\n|\s*$))', block, re.IGNORECASE)
            if m_fondo: notas['fondo'] = _parse_notas_list(m_fondo.group(1))
        
        # 2. Extract specific phrases if structured fails or is partial
        if not notas['salida'] and not notas['corazon'] and not notas['fondo']:
            m_familia = re.search(r'familia olfativa[\s:]*(.+?)(?:\.|\n|$)', text, re.IGNORECASE)
            if m_familia:
                notas['general'].extend(_parse_notas_list(m_familia.group(1)))
            m_notas = re.search(r'notas principales[\s:]*(.+?)(?:\.|\n|$)', text, re.IGNORECASE)
            if m_notas:
                notas['general'].extend(_parse_notas_list(m_notas.group(1)))
        
        # 3. Last resort: scan the text for ANY keyword from our dictionary
        if not any(notas.values()):
            found = set()
            text_lower = text.lower() + " " + r['nombre'].lower()
            for fam, keywords in AROMA_FAMILIES.items():
                for kw in keywords:
                    if kw in text_lower:
                        found.add(kw)
            if found:
                notas['general'] = list(found)

        if any(notas.values()):
            _query(db, f'UPDATE {table} SET aromas = %s WHERE sku = %s',
                   [json.dumps(notas, ensure_ascii=False), r['sku']])
            extracted += 1

    return extracted


def _parse_notas_list(raw):
    """Parse a raw notes string like 'bergamota, piña y ámbar' into a list."""
    raw = raw.strip().rstrip('.,;:')
    # Remove trailing connectors
    raw = re.sub(r'\s+(?:y|e)\s*$', '', raw)
    items = [x.strip().lower() for x in re.split(r'\s*,\s*|\s+y\s+|\s+e\s+', raw) if x.strip()]
    # Filter out noise
    return [i for i in items if len(i) > 2 and not i.startswith('nota')]



@app.route('/admin/migrar-db')
@require_login
def migrar_db():
    db = get_db()
    c = db.cursor(dictionary=True)
    c.execute("SELECT * FROM cosmetic_products")
    rows = c.fetchall()
    
    today = date.today().strftime('%Y-%m-%d')
    _query(db, 'INSERT IGNORE INTO imports (import_date) VALUES (%s)', [today])
    
    added = 0
    for r in rows:
        aroma_family = None
        if r['aromas']:
            try:
                notas = json.loads(r['aromas'])
                fams = _classify_family(notas)
                if fams:
                    aroma_family = fams[0]
            except Exception:
                pass
                
        if not aroma_family:
            fams = _classify_family({'general': [r['nombre']]})
            if fams:
                aroma_family = fams[0]

        linea = r['nombre'].split()[0] if r['nombre'] else ''
        
        _query(db, '''INSERT INTO products (sku, nombre, linea, aroma_family, is_active, stock)
                      VALUES (%s, %s, %s, %s, 1, 10)
                      ON DUPLICATE KEY UPDATE 
                      nombre=VALUES(nombre), aroma_family=VALUES(aroma_family)''',
               [r['sku'], r['nombre'], linea, aroma_family])
               
        precio = r['precio_retail'] if r['precio_retail'] else 0
        if precio > 0:
            exists = _query(db, 'SELECT id FROM prices WHERE sku = %s AND import_date = %s', [r['sku'], today]).fetchone()
            if not exists:
                _query(db, 'INSERT INTO prices (sku, import_date, precio) VALUES (%s, %s, %s)', [r['sku'], today, precio])
            else:
                _query(db, 'UPDATE prices SET precio = %s WHERE sku = %s AND import_date = %s', [precio, r['sku'], today])
        added += 1

    db.commit()
    flash(f'¡Migración completada! Se aplicaron {added} productos de cosmetic.cl a tu inventario real.', 'success')
    return redirect(url_for('admin_productos'))

@app.route('/sync-cosmetic')
@require_login
def sync_cosmetic():
    db = get_db()
    try:
        added, error = _sync_shopify(db, 'cosmetic_products', 'https://cosmetic.cl')
        extracted = 0
        if not error:
            extracted = _extract_aromas(db, 'cosmetic_products')
    finally:
        db.close()
    if error:
        flash(f'Error sync cosmetic.cl: {error}', 'error')
    elif added == 0:
        flash('No se encontraron productos en cosmetic.cl', 'warning')
    else:
        flash(f'Sincronizados {added} productos, {extracted} con aromas de cosmetic.cl', 'success')
    return redirect(url_for('retail'))


@app.route('/sync-silk')
@require_login
def sync_silk():
    db = get_db()
    try:
        added, error = _sync_shopify(db, 'silk_products', 'https://silkperfumes.cl')
        if not error:
            _match_silk_by_name(db)
    finally:
        db.close()
    if error:
        flash(f'Error sync silkperfumes.cl: {error}', 'error')
    elif added == 0:
        flash('No se encontraron productos en silkperfumes.cl', 'warning')
    else:
        flash(f'Sincronizados {added} productos de silkperfumes.cl', 'success')
    return redirect(url_for('retail'))


@app.route('/sync-multimarca')
@require_login
def sync_multimarca():
    db = get_db()
    try:
        added, error = _sync_shopify(db, 'multimarca_products', 'https://multimarcasperfumes.cl')
        matched = 0
        extracted = 0
        if not error:
            matched = _match_by_name(db, 'multimarca_matches', 'multimarca_products', 'sku_mm')
            extracted = _extract_aromas(db, 'multimarca_products')
    finally:
        db.close()
    if error:
        flash(f'Error sync multimarcasperfumes.cl: {error}', 'error')
    elif added == 0:
        flash('No se encontraron productos en multimarcasperfumes.cl', 'warning')
    else:
        flash(f'Sincronizados {added} productos, {matched} matcheados, {extracted} aromas de multimarcasperfumes.cl', 'success')
    return redirect(url_for('retail'))


def _normalize(n):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', '', n.lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u'))).strip()


def _match_by_name(db, match_table, source_table, sku_col):
    """Generic fuzzy name matching between wholesale products and a retail source."""
    autocommit = db.autocommit
    db.autocommit = False
    try:
        _query(db, f'DELETE FROM {match_table}')
        wholesale = _query(db, 'SELECT sku, nombre, linea FROM products').fetchall()
        source = _query(db, f'SELECT sku, nombre FROM {source_table}').fetchall()

        source_by_token = {}
        for s in source:
            sn = _normalize(s['nombre'])
            tokens = sn.split()
            if tokens:
                k1 = tokens[0]
                source_by_token.setdefault(k1, []).append((sn, s['sku']))
                if len(tokens) >= 2:
                    k2 = f'{tokens[0]} {tokens[1]}'
                    source_by_token.setdefault(k2, []).append((sn, s['sku']))

        matched = 0
        for w in wholesale:
            w_norm = _normalize(w['nombre'])
            w_brand = _normalize(w['linea'])
            tokens = w_norm.split()
            w_key1 = tokens[0] if tokens else ''
            w_key2 = f'{tokens[0]} {tokens[1]}' if len(tokens) >= 2 else w_key1

            seen = set()
            candidates = []
            for key in [w_brand, w_key1, w_key2]:
                for sn, ssku in source_by_token.get(key, []):
                    if ssku not in seen:
                        seen.add(ssku)
                        candidates.append((sn, ssku))

            if w_brand and len(w_brand) > 3:
                for s in source:
                    if s['sku'] in seen:
                        continue
                    sn = _normalize(s['nombre'])
                    if w_brand in sn:
                        seen.add(s['sku'])
                        candidates.append((sn, s['sku']))
                        if len(candidates) > 200:
                            break

            if not candidates:
                continue

            best_score = 0
            best_sku = None
            for sn, ssku in candidates:
                if abs(len(w_norm) - len(sn)) > len(w_norm) * 0.6:
                    continue
                score = SequenceMatcher(None, w_norm, sn).ratio()
                if score > best_score:
                    best_score = score
                    best_sku = ssku

            if best_score >= 0.70:
                _query(db, f'INSERT INTO {match_table} (sku_wholesale, {sku_col}, confidence) VALUES (%s, %s, %s)',
                       [w['sku'], best_sku, round(best_score, 3)])
                matched += 1

        db.commit()
        return matched
    except Exception:
        db.rollback()
        raise
    finally:
        db.autocommit = autocommit


def _match_silk_by_name(db):
    return _match_by_name(db, 'silk_matches', 'silk_products', 'sku_silk')


@app.route('/retail')
@require_login
def retail():
    query = request.args.get('q', '').strip()
    linea = request.args.get('linea', '').strip()
    sort = request.args.get('sort', 'diff_cosmetic_desc')
    margen_obj = request.args.get('margen', '30')
    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = 50
    try:
        margen_pct = float(margen_obj)
    except ValueError:
        margen_pct = 30
    db = get_db()
    latest = _query(db, 'SELECT MAX(import_date) as max_date FROM imports').fetchone()['max_date']
    lineas = [r['linea'] for r in _query(db, 'SELECT DISTINCT linea FROM products ORDER BY linea')]
    synced = _query(db, 'SELECT COUNT(*) as cnt, MAX(last_synced) as last FROM cosmetic_products').fetchone()

    sql = '''SELECT p.sku, p.nombre, p.linea, p.genero, p.formato,
                    pr.precio as precio_mayorista,
                    cp.precio_retail as precio_cosmetic, cp.precio_ref as ref_cosmetic,
                    cp.imagen as img_cosmetic, cp.url as url_cosmetic,
                    sp.precio_retail as precio_silk, sp.precio_ref as ref_silk,
                    sp.imagen as img_silk, sp.url as url_silk,
                    mp.precio_retail as precio_mm, mp.precio_ref as ref_mm,
                    mp.imagen as img_mm, mp.url as url_mm,
                    (cp.precio_retail - pr.precio) as diff_cosmetic,
                    (sp.precio_retail - pr.precio) as diff_silk,
                    (mp.precio_retail - pr.precio) as diff_mm,
                    CASE WHEN cp.precio_retail > 0 THEN ROUND((cp.precio_retail - pr.precio) * 100.0 / cp.precio_retail, 1) END as margen_cosmetic,
                    CASE WHEN sp.precio_retail > 0 THEN ROUND((sp.precio_retail - pr.precio) * 100.0 / sp.precio_retail, 1) END as margen_silk,
                    CASE WHEN mp.precio_retail > 0 THEN ROUND((mp.precio_retail - pr.precio) * 100.0 / mp.precio_retail, 1) END as margen_mm
             FROM products p
             JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
             JOIN cosmetic_products cp ON p.sku = cp.sku
             LEFT JOIN silk_matches sm ON p.sku = sm.sku_wholesale
             LEFT JOIN silk_products sp ON sm.sku_silk = sp.sku
             LEFT JOIN multimarca_matches mm ON p.sku = mm.sku_wholesale
             LEFT JOIN multimarca_products mp ON mm.sku_mm = mp.sku'''
    params = [latest] if latest else [None]
    conditions = []
    if query:
        conditions.append('(p.nombre LIKE %s OR p.sku LIKE %s OR p.linea LIKE %s)')
        like = f'%{query}%'
        params.extend([like, like, like])
    if linea:
        conditions.append('p.linea = %s')
        params.append(linea)
    if conditions:
        sql += ' AND ' + ' AND '.join(conditions)

    sort_map = {
        'diff_cosmetic_desc': '(cp.precio_retail - pr.precio) DESC',
        'diff_silk_desc': '(sp.precio_retail - pr.precio) DESC',
        'margen_cosmetic': 'CASE WHEN cp.precio_retail > 0 THEN (cp.precio_retail - pr.precio) * 100.0 / cp.precio_retail ELSE 0 END DESC',
        'margen_silk': 'CASE WHEN sp.precio_retail > 0 THEN (sp.precio_retail - pr.precio) * 100.0 / sp.precio_retail ELSE 0 END DESC',
        'linea': 'p.linea, p.nombre',
    }
    sql += f' ORDER BY {sort_map.get(sort, sort_map["diff_cosmetic_desc"])}'

    # Count total
    total = _query(db, f'SELECT COUNT(*) AS cnt FROM ({sql}) AS t', params).fetchone()['cnt']
    offset = (page - 1) * per_page
    products = _query(db, f'{sql} LIMIT %s OFFSET %s', params + [per_page, offset]).fetchall()
    total_pages = (total + per_page - 1) // per_page if total else 1
    db.close()

    con_cosmetic = sum(1 for p in products if p['precio_cosmetic'] > 0)
    con_silk = sum(1 for p in products if p['precio_silk'] and p['precio_silk'] > 0)
    con_mm = sum(1 for p in products if p['precio_mm'] and p['precio_mm'] > 0)
    return render_template('retail.html', products=products, query=query,
                          linea=linea, lineas=lineas, latest=latest,
                          synced=synced, total=total,
                          con_cosmetic=con_cosmetic, con_silk=con_silk, con_mm=con_mm,
                          sort=sort, margen_obj=margen_obj, margen_pct=margen_pct,
                          page=page, per_page=per_page, total_pages=total_pages)


@app.route('/retail/export')
@require_login
def retail_export():
    skus = request.args.get('skus', '').split(',')
    if not skus or skus == ['']:
        flash('Seleccioná al menos un producto', 'error')
        return redirect(url_for('retail'))
    db = get_db()
    latest = _query(db, 'SELECT MAX(import_date) as max_date FROM imports').fetchone()['max_date']
    placeholders = ','.join(['%s'] * len(skus))
    rows = _query(db, f'''
        SELECT p.sku, p.nombre, p.linea, p.genero, p.formato,
               pr.precio as mayorista,
               cp.precio_retail as cosmetic, cp.url as url_cosmetic, cp.imagen as img_cosmetic,
               sp.precio_retail as silk, sp.url as url_silk, sp.imagen as img_silk
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        JOIN cosmetic_products cp ON p.sku = cp.sku
        LEFT JOIN silk_matches sm ON p.sku = sm.sku_wholesale
        LEFT JOIN silk_products sp ON sm.sku_silk = sp.sku
        WHERE p.sku IN ({placeholders})
        GROUP BY p.sku, p.nombre, p.linea, p.genero, p.formato, pr.precio,
                 cp.precio_retail, cp.url, cp.imagen, sp.precio_retail, sp.url, sp.imagen
        ORDER BY p.linea, p.nombre
    ''', [latest] + skus).fetchall()
    db.close()

    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['SKU', 'Marca', 'Nombre', 'Genero', 'Formato', 'Mayorista',
                     'cosmetic.cl', 'silkperfumes.cl', 'URL cosmetic', 'URL silk',
                     'Imagen cosmetic', 'Imagen silk'])
    for r in rows:
        writer.writerow([
            r['sku'], r['linea'], r['nombre'], r['genero'], r['formato'],
            r['mayorista'], r['cosmetic'] or '', r['silk'] or '',
            r['url_cosmetic'] or '', r['url_silk'] or '',
            r['img_cosmetic'] or '', r['img_silk'] or ''
        ])
    output.seek(0)
    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = 'attachment; filename=comparativa.csv'
    return resp


@app.route('/catalogo/pdf')
def catalogo_pdf():
    skus = request.args.get('skus', '').split(',')
    if not skus or skus == ['']:
        return 'Seleccioná al menos un producto', 400

    try:
        margen_pct = float(request.args.get('margen', '30'))
    except ValueError:
        margen_pct = 30

    db = get_db()
    latest = _query(db, 'SELECT MAX(import_date) as max_date FROM imports').fetchone()['max_date']
    placeholders = ','.join(['%s'] * len(skus))
    rows = _query(db, f'''
        SELECT p.sku, p.nombre, p.linea, p.genero, p.formato,
               cp.imagen, cp.url, cp.aromas, cp.precio_retail,
               MAX(pr.precio) as precio_mayorista
        FROM products p
        JOIN cosmetic_products cp ON p.sku = cp.sku
        LEFT JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        WHERE p.sku IN ({placeholders})
        GROUP BY p.sku, p.nombre, p.linea, p.genero, p.formato,
                 cp.imagen, cp.url, cp.aromas, cp.precio_retail
        ORDER BY p.linea, p.nombre
    ''', [latest] + skus).fetchall()
    db.close()

    if not rows:
        return 'No se encontraron productos con esos SKU', 404

    for p in rows:
        p['notas'] = {}
        if p['aromas']:
            try:
                p['notas'] = json.loads(p['aromas'])
            except (json.JSONDecodeError, TypeError):
                pass

    pdf_bytes = _generate_catalogo_pdf(rows, margen_pct)
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = 'attachment; filename=catalogo_perfumes.pdf'
    return resp


def _generate_catalogo_pdf(products, margen_pct=30):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, Image as RLImage)
    import urllib.request as req

    buf = BytesIO()
    page_w, page_h = landscape(A4)
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                           leftMargin=10*mm, rightMargin=10*mm,
                           topMargin=10*mm, bottomMargin=10*mm)

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('CatTitle', parent=styles['Heading1'],
                                 fontSize=14, spaceAfter=4*mm, textColor=HexColor('#1e293b'))
    style_cell = ParagraphStyle('CatCell', parent=styles['Normal'],
                                fontSize=8, leading=10, textColor=HexColor('#334155'))
    style_cell_bold = ParagraphStyle('CatCellBold', parent=style_cell,
                                     fontSize=8, leading=10, textColor=HexColor('#1e293b'))
    style_header = ParagraphStyle('CatHeader', parent=styles['Normal'],
                                  fontSize=8, leading=10, textColor=HexColor('#ffffff'))

    header_bg = HexColor('#6366f1')
    row_even = HexColor('#f8fafc')
    row_odd = HexColor('#ffffff')
    border_color = HexColor('#e2e8f0')

    story = [Paragraph('Catálogo de Perfumes', style_title),
             Paragraph(f'{len(products)} productos', styles['Normal']),
             Spacer(1, 4*mm)]

    # ponytail: parallel pre-fetch — secuencial era O(n × timeout)
    img_size = 18*mm
    products_sorted = sorted(products, key=lambda p: (p['linea'] or '', p['nombre'] or ''))

    # --- parallel image fetch ---
    unique_urls = {p.get('imagen') for p in products_sorted if p.get('imagen')}
    _image_cache = {}

    def _fetch_one(url):
        try:
            r = req.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with req.urlopen(r, timeout=3) as resp:
                return url, resp.read()
        except Exception:
            return url, None

    with ThreadPoolExecutor(max_workers=20) as ex:
        futures = {ex.submit(_fetch_one, url): url for url in unique_urls}
        for f in as_completed(futures):
            url, data = f.result()
            _image_cache[url] = data
    # --- end parallel fetch ---

    # 7 columns: Imagen, Marca, Nombre, Salida, Corazón, Fondo, Precio
    aroma_w = 50*mm
    col_widths = [img_size + 4*mm, 24*mm, 52*mm, aroma_w, aroma_w, aroma_w, 22*mm]

    header = [
        Paragraph('<b>Imagen</b>', style_header),
        Paragraph('<b>Marca</b>', style_header),
        Paragraph('<b>Nombre</b>', style_header),
        Paragraph('<b>Salida</b>', style_header),
        Paragraph('<b>Corazón</b>', style_header),
        Paragraph('<b>Fondo</b>', style_header),
        Paragraph('<b>Precio</b>', style_header),
    ]

    rows = [header]
    for i, p in enumerate(products_sorted):
        img_data = _image_cache.get(p.get('imagen'))
        if img_data:
            try:
                img = RLImage(BytesIO(img_data), width=img_size, height=img_size)
            except Exception:
                img = Paragraph('—', style_cell)
        else:
            img = Paragraph('—', style_cell)

        brand = Paragraph(p['linea'] or '', style_cell_bold)
        name = Paragraph(p['nombre'] or '', style_cell)
        notas = p.get('notas', {})
        salida = Paragraph(', '.join(a.capitalize() for a in notas.get('salida', [])) or '—', style_cell)
        corazon = Paragraph(', '.join(a.capitalize() for a in notas.get('corazon', [])) or '—', style_cell)
        fondo = Paragraph(', '.join(a.capitalize() for a in notas.get('fondo', [])) or '—', style_cell)

        retail = p.get('precio_retail')
        if retail and retail > 0:
            precio_str = f"${'{:,}'.format(retail).replace(',', '.')}"
        else:
            precio_str = '—'
        precio = Paragraph(precio_str, style_cell_bold)

        row = [img, brand, name, salida, corazon, fondo, precio]
        rows.append(row)

    # Build table
    t = Table(rows, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#ffffff')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        # Grid lines
        ('GRID', (0, 0), (-1, -1), 0.3, border_color),
        ('LINEBELOW', (0, 0), (-1, 0), 0.6, HexColor('#4f46e5')),
        # Alignment
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # image centered
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    # Alternating row colors (skip header row 0)
    for i in range(1, len(rows)):
        bg = row_even if i % 2 == 0 else row_odd
        style_cmds.append(('BACKGROUND', (0, i), (-1, i), bg))

    t.setStyle(TableStyle(style_cmds))
    story.append(t)

    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}',
                           styles['Normal']))

    doc.build(story)
    return buf.getvalue()


@app.route('/estudio')
@require_login
def estudio():
    db = get_db()
    latest = _query(db, 'SELECT MAX(import_date) as max_date FROM imports').fetchone()['max_date']
    total_wholesale = _query(db, 'SELECT COUNT(*) as cnt FROM products').fetchone()['cnt']
    total_cosmetic = _query(db, 'SELECT COUNT(*) as cnt FROM cosmetic_products').fetchone()['cnt']
    total_silk = _query(db, 'SELECT COUNT(*) as cnt FROM silk_products').fetchone()['cnt']
    total_mm = _query(db, 'SELECT COUNT(*) as cnt FROM multimarca_products').fetchone()['cnt']
    matched_cosmetic = _query(db, 'SELECT COUNT(*) as cnt FROM products p JOIN cosmetic_products cp ON p.sku = cp.sku').fetchone()['cnt']
    matched_silk = _query(db, 'SELECT COUNT(*) as cnt FROM silk_matches').fetchone()['cnt']
    matched_mm = _query(db, 'SELECT COUNT(*) as cnt FROM multimarca_matches').fetchone()['cnt']
    match_rate_cosmetic = round(matched_cosmetic * 100.0 / total_wholesale, 1) if total_wholesale else 0
    match_rate_silk = round(matched_silk * 100.0 / total_wholesale, 1) if total_wholesale else 0
    match_rate_mm = round(matched_mm * 100.0 / total_wholesale, 1) if total_wholesale else 0

    avg_gap = _query(db, '''
        SELECT ROUND(AVG(cp.precio_retail - pr.precio)) as avg_diff,
               ROUND(AVG((cp.precio_retail - pr.precio) * 100.0 / cp.precio_retail), 1) as avg_margin
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        JOIN cosmetic_products cp ON p.sku = cp.sku
        WHERE cp.precio_retail > 0
    ''', [latest]).fetchone()

    avg_gap_mm = _query(db, '''
        SELECT ROUND(AVG(mp.precio_retail - pr.precio)) as avg_diff,
               ROUND(AVG((mp.precio_retail - pr.precio) * 100.0 / mp.precio_retail), 1) as avg_margin
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        JOIN multimarca_matches mm ON p.sku = mm.sku_wholesale
        JOIN multimarca_products mp ON mm.sku_mm = mp.sku
        WHERE mp.precio_retail > 0
    ''', [latest]).fetchone()

    brand_stats = _query(db, '''
        SELECT p.linea, COUNT(*) as n,
               ROUND(AVG(pr.precio)) as avg_costo,
               ROUND(AVG(cp.precio_retail)) as avg_retail,
               ROUND(AVG((cp.precio_retail - pr.precio) * 100.0 / cp.precio_retail), 1) as avg_margin
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        JOIN cosmetic_products cp ON p.sku = cp.sku
        WHERE cp.precio_retail > 0
        GROUP BY p.linea HAVING n >= 3
        ORDER BY avg_margin DESC LIMIT 20
    ''', [latest]).fetchall()

    brand_stats_mm = _query(db, '''
        SELECT p.linea, COUNT(*) as n,
               ROUND(AVG(pr.precio)) as avg_costo,
               ROUND(AVG(mp.precio_retail)) as avg_retail,
               ROUND(AVG((mp.precio_retail - pr.precio) * 100.0 / mp.precio_retail), 1) as avg_margin
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        JOIN multimarca_matches mm ON p.sku = mm.sku_wholesale
        JOIN multimarca_products mp ON mm.sku_mm = mp.sku
        WHERE mp.precio_retail > 0
        GROUP BY p.linea HAVING n >= 3
        ORDER BY avg_margin DESC LIMIT 20
    ''', [latest]).fetchall()

    opportunities = _query(db, '''
        SELECT p.sku, p.nombre, p.linea, pr.precio as costo,
               cp.precio_retail, cp.url, cp.imagen,
               (cp.precio_retail - pr.precio) as diff,
               ROUND((cp.precio_retail - pr.precio) * 100.0 / cp.precio_retail, 1) as margen
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        JOIN cosmetic_products cp ON p.sku = cp.sku
        WHERE cp.precio_retail > 0
        ORDER BY margen DESC LIMIT 30
    ''', [latest]).fetchall()

    opportunities_mm = _query(db, '''
        SELECT p.sku, p.nombre, p.linea, pr.precio as costo,
               mp.precio_retail, mp.url, mp.imagen,
               (mp.precio_retail - pr.precio) as diff,
               ROUND((mp.precio_retail - pr.precio) * 100.0 / mp.precio_retail, 1) as margen
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        JOIN multimarca_matches mm ON p.sku = mm.sku_wholesale
        JOIN multimarca_products mp ON mm.sku_mm = mp.sku
        WHERE mp.precio_retail > 0
        ORDER BY margen DESC LIMIT 30
    ''', [latest]).fetchall()

    top_cost = _query(db, '''
        SELECT p.sku, p.nombre, p.linea, pr.precio, cp.imagen
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        JOIN cosmetic_products cp ON p.sku = cp.sku
        ORDER BY pr.precio DESC LIMIT 10
    ''', [latest]).fetchall()

    db.close()
    return render_template('estudio.html',
                          total_wholesale=total_wholesale,
                          total_cosmetic=total_cosmetic, total_silk=total_silk, total_mm=total_mm,
                          matched_cosmetic=matched_cosmetic, matched_silk=matched_silk, matched_mm=matched_mm,
                          match_rate_cosmetic=match_rate_cosmetic, match_rate_silk=match_rate_silk,
                          match_rate_mm=match_rate_mm,
                          avg_gap=avg_gap, avg_gap_mm=avg_gap_mm,
                          brand_stats=brand_stats, brand_stats_mm=brand_stats_mm,
                          opportunities=opportunities, opportunities_mm=opportunities_mm,
                          top_cost=top_cost, latest=latest)


@app.route('/export/xlsx')
@require_login
def export_xlsx():
    """Export selected products as XLSX."""
    skus = request.args.get('skus', '').split(',')
    if not skus or skus == ['']:
        flash('Seleccioná al menos un producto', 'error')
        return redirect(url_for('admin_productos'))

    db = get_db()
    latest = _query(db, 'SELECT MAX(import_date) as max_date FROM imports').fetchone()['max_date']
    placeholders = ','.join(['%s'] * len(skus))
    rows = _query(db, f'''
        SELECT p.sku, p.nombre, p.linea, p.ean, p.genero, p.formato, pr.precio
        FROM products p
        JOIN prices pr ON p.sku = pr.sku AND pr.import_date = %s
        WHERE p.sku IN ({placeholders})
        ORDER BY p.linea, p.nombre
    ''', [latest] + skus).fetchall()
    db.close()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Lista de Precios'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_align = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    price_font = Font(bold=True)

    headers = ['SKU', 'Marca', 'Nombre', 'EAN', 'Género', 'Formato', 'Precio']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(rows, 2):
        vals = [r['sku'], r['linea'], r['nombre'], r['ean'], r['genero'], r['formato'], r['precio']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            if col == 7:
                cell.font = price_font
                cell.number_format = '$#,##0'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=lista_precios.xlsx'
    return resp


@app.route('/pedidos')
@require_login
def pedidos():
    db = get_db()
    status_filter = request.args.get('status', '').strip()
    sql = '''SELECT o.*, cl.margen_pct
             FROM orders o
             LEFT JOIN catalog_links cl ON o.link_token = cl.token'''
    params = []
    if status_filter:
        sql += ' WHERE o.status = %s'
        params.append(status_filter)
    sql += ' ORDER BY o.created_at DESC LIMIT 200'
    orders_list = _query(db, sql, params).fetchall()

    for o in orders_list:
        try:
            o['items'] = json.loads(o['items_json'])
        except (json.JSONDecodeError, TypeError):
            o['items'] = []

    counts = _query(db, '''
        SELECT status, COUNT(*) as cnt FROM orders GROUP BY status
    ''').fetchall()
    db.close()
    return render_template('pedidos.html', orders=orders_list, counts=counts, status_filter=status_filter)


@app.route('/pedidos/<int:order_id>/status', methods=['POST'])
@require_login
def pedido_status(order_id):
    nuevo_status = request.form.get('status', '').strip()
    if nuevo_status not in ('nuevo', 'procesando', 'completado', 'cancelado'):
        flash('Estado inválido', 'error')
        return redirect(url_for('pedidos'))
    db = get_db()
    _query(db, 'UPDATE orders SET status = %s WHERE id = %s', [nuevo_status, order_id])
    db.close()
    flash('Estado actualizado', 'success')
    return redirect(url_for('pedidos'))


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit('5 per minute')
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        db = get_db()
        user = _query(db, 'SELECT * FROM users WHERE username = %s', [username]).fetchone()
        db.close()
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            flash(f'Bienvenido, {user["username"]}', 'success')
            return redirect(url_for('admin_productos'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada', 'success')
    return redirect(url_for('login'))


# ==================== TIENDA (ecommerce público) ====================

_HAS_IMG_SQL = "(COALESCE(NULLIF(cp.imagen, ''), NULLIF(mp.imagen, ''), NULLIF(sp.imagen, '')) IS NOT NULL)"


def _tienda_sql():
    """Base SQL for storefront products. First param must be the margin factor."""
    return '''SELECT p.sku, p.nombre, p.linea, p.genero, p.formato, p.aroma_family, p.stock,
                     pr.precio AS precio_mayorista,
                     (ROUND((pr.precio * %s + 10) / 1000) * 1000 - 10) AS precio_venta,
                     COALESCE(NULLIF(cp.imagen, ''), NULLIF(mp.imagen, ''), NULLIF(sp.imagen, '')) AS imagen,
                     COALESCE(cp.aromas, mp.aromas) AS aromas
              FROM products p
              JOIN prices pr ON p.sku = pr.sku
                   AND pr.import_date = (SELECT MAX(import_date) FROM imports)
              LEFT JOIN cosmetic_products cp ON cp.sku = p.sku
              LEFT JOIN multimarca_matches mmk ON mmk.sku_wholesale = p.sku
              LEFT JOIN multimarca_products mp ON mp.sku = mmk.sku_mm
              LEFT JOIN silk_matches smk ON smk.sku_wholesale = p.sku
              LEFT JOIN silk_products sp ON sp.sku = smk.sku_silk
              WHERE p.is_active = 1 AND p.stock > 0'''


def _enrich_products(products):
    """Normalize gender and parse fragrance notes for display."""
    for p in products:
        p['genero_norm'] = _norm_genero(p.get('genero'))
        p['notas'] = {}
        if p.get('aromas'):
            try:
                p['notas'] = json.loads(p['aromas'])
            except (json.JSONDecodeError, TypeError):
                p['notas'] = {}
        familias = _classify_family(p['notas']) if p['notas'] else []
        if not familias and p.get('aroma_family'):
            familias = [p['aroma_family']]
        p['familias'] = familias
    return products


def _get_margen(db):
    try:
        return float(get_setting(db, 'margen_pct') or 30)
    except (TypeError, ValueError):
        return 30.0


@app.route('/')
def tienda_home():
    db = get_db()
    factor = 1 + _get_margen(db) / 100.0

    destacados = _query(db, _tienda_sql() + f'''
        ORDER BY {_HAS_IMG_SQL} DESC, pr.precio DESC
        LIMIT 8''', [factor]).fetchall()
    _enrich_products(destacados)

    total = _query(db, 'SELECT COUNT(*) AS cnt FROM products p WHERE p.is_active = 1 AND p.stock > 0').fetchone()['cnt']

    def _count_genero(genero_norm):
        conds, pats = _genero_sql_conds(genero_norm)
        if not conds:
            return 0
        sql = 'SELECT COUNT(*) AS cnt FROM products p WHERE p.is_active = 1 AND p.stock > 0 AND (' + ' OR '.join(conds) + ')'
        return _query(db, sql, pats).fetchone()['cnt']

    n_hombre = _count_genero('Hombre')
    n_mujer = _count_genero('Mujer')
    n_unisex = max(0, total - n_hombre - n_mujer)

    marcas = _query(db, '''SELECT p.linea, COUNT(*) AS n,
                                  COALESCE(NULLIF(cp.imagen, ''), NULLIF(mp.imagen, ''), NULLIF(sp.imagen, '')) AS imagen
                           FROM products p
                           LEFT JOIN cosmetic_products cp ON cp.sku = p.sku
                           LEFT JOIN multimarca_matches mmk ON mmk.sku_wholesale = p.sku
                           LEFT JOIN multimarca_products mp ON mp.sku = mmk.sku_mm
                           LEFT JOIN silk_matches smk ON smk.sku_wholesale = p.sku
                           LEFT JOIN silk_products sp ON sp.sku = smk.sku_silk
                           WHERE p.is_active = 1 AND p.stock > 0
                           GROUP BY p.linea 
                           ORDER BY n DESC LIMIT 16''').fetchall()

    familias = _query(db, '''SELECT p.aroma_family AS fam, COUNT(*) AS n FROM products p
                             WHERE p.is_active = 1 AND p.stock > 0 AND p.aroma_family IS NOT NULL
                             GROUP BY p.aroma_family ORDER BY n DESC''').fetchall()
    db.close()

    return render_template('tienda/home.html', destacados=destacados,
                           total=total, n_hombre=n_hombre, n_mujer=n_mujer,
                           n_unisex=n_unisex, marcas=marcas, familias=familias)


def _catalogo_response(preset_genero=None):
    q = request.args.get('q', '').strip()
    genero_raw = preset_genero or (request.args.get('genero', '').strip() or None)
    genero = _norm_genero(genero_raw) if genero_raw else None
    linea = request.args.get('linea', '').strip()
    familia = request.args.get('familia', '').strip()
    sort = request.args.get('sort', 'destacados').strip()
    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = 24

    def _int_arg(name):
        try:
            v = int(float(request.args.get(name, '') or 0))
            return v if v > 0 else None
        except ValueError:
            return None

    pmin = _int_arg('pmin')
    pmax = _int_arg('pmax')

    db = get_db()
    factor = 1 + _get_margen(db) / 100.0

    lineas = [r['linea'] for r in _query(db, '''SELECT DISTINCT p.linea FROM products p
                                                WHERE p.is_active = 1 AND p.stock > 0
                                                ORDER BY p.linea''').fetchall()]

    sql = _tienda_sql()
    params = [factor]
    conditions = []
    if q:
        conditions.append('(p.nombre LIKE %s OR p.linea LIKE %s OR p.sku LIKE %s)')
        like = f'%{q}%'
        params.extend([like, like, like])
    if linea:
        conditions.append('p.linea = %s')
        params.append(linea)
    if genero:
        conds, pats = _genero_sql_conds(genero)
        if conds:
            conditions.append('(' + ' OR '.join(conds) + ')')
            params.extend(pats)
    if familia and familia in AROMA_FAMILIES:
        fam_conds = ['p.aroma_family = %s']
        params.append(familia)
        for kw in AROMA_FAMILIES[familia]:
            fam_conds.append('COALESCE(cp.aromas, mp.aromas) LIKE %s')
            params.append(f'%{kw}%')
        conditions.append('(' + ' OR '.join(fam_conds) + ')')
    if pmin:
        conditions.append('(ROUND((pr.precio * %s + 10) / 1000) * 1000 - 10) >= %s')
        params.append(factor)
        params.append(pmin)
    if pmax:
        conditions.append('(ROUND((pr.precio * %s + 10) / 1000) * 1000 - 10) <= %s')
        params.append(factor)
        params.append(pmax)
    if conditions:
        sql += ' AND ' + ' AND '.join(conditions)

    sort_map = {
        'destacados': f'{_HAS_IMG_SQL} DESC, pr.precio DESC',
        'precio_asc': 'pr.precio ASC',
        'precio_desc': 'pr.precio DESC',
        'nombre': 'p.nombre ASC',
        'marca': 'p.linea ASC, p.nombre ASC',
    }
    sql += f' ORDER BY {sort_map.get(sort, sort_map["destacados"])}'

    total = _query(db, f'SELECT COUNT(*) AS cnt FROM ({sql}) AS t', params).fetchone()['cnt']
    offset = (page - 1) * per_page
    products = _query(db, f'{sql} LIMIT %s OFFSET %s', params + [per_page, offset]).fetchall()
    total_pages = (total + per_page - 1) // per_page if total else 1
    _enrich_products(products)

    bounds = _query(db, '''SELECT (ROUND((MIN(pr.precio) * %s + 10) / 1000) * 1000 - 10) AS pmin, (ROUND((MAX(pr.precio) * %s + 10) / 1000) * 1000 - 10) AS pmax
                           FROM products p JOIN prices pr ON pr.sku = p.sku
                           WHERE p.is_active = 1 AND p.stock > 0
                           AND pr.import_date = (SELECT MAX(import_date) FROM imports)''',
                    [factor, factor]).fetchone()
    db.close()

    return render_template('tienda/catalogo.html', products=products, q=q,
                           genero=genero, linea=linea, familia=familia, sort=sort,
                           pmin=pmin, pmax=pmax,
                           bounds_min=int(bounds['pmin'] or 0) if bounds['pmin'] else 0,
                           bounds_max=int(bounds['pmax'] or 0) if bounds['pmax'] else 0,
                           lineas=lineas, familias=sorted(AROMA_FAMILIES.keys()),
                           total=total, page=page, per_page=per_page,
                           total_pages=total_pages, preset=preset_genero or '')


@app.route('/perfumes')
def tienda_catalogo():
    return _catalogo_response(None)


@app.route('/hombre')
def tienda_hombre():
    return _catalogo_response('Hombre')


@app.route('/mujer')
def tienda_mujer():
    return _catalogo_response('Mujer')


@app.route('/producto/<sku>')
def tienda_producto(sku):
    db = get_db()
    factor = 1 + _get_margen(db) / 100.0
    product = _query(db, _tienda_sql() + ' AND p.sku = %s', [factor, sku]).fetchone()
    if not product:
        db.close()
        return render_template('tienda/404.html'), 404
    _enrich_products([product])

    relacionados = _query(db, _tienda_sql() + ' AND p.sku != %s AND p.linea = %s '
                        + f'ORDER BY {_HAS_IMG_SQL} DESC, pr.precio DESC LIMIT 4',
                        [factor, sku, product['linea']]).fetchall()
    if len(relacionados) < 4:
        conds, pats = _genero_sql_conds(product['genero_norm'])
        extra_sql = _tienda_sql() + ' AND p.sku != %s AND p.linea != %s'
        extra_params = [factor, sku, product['linea']]
        if conds:
            extra_sql += ' AND (' + ' OR '.join(conds) + ')'
            extra_params.extend(pats)
        extra_sql += f' ORDER BY {_HAS_IMG_SQL} DESC, pr.precio DESC LIMIT %s'
        extra_params.append(4 - len(relacionados))
        relacionados.extend(_query(db, extra_sql, extra_params).fetchall())
    _enrich_products(relacionados)
    db.close()

    return render_template('tienda/producto.html', p=product, relacionados=relacionados)


@app.route('/api/pedido', methods=['POST'])
@csrf.exempt
def api_pedido():
    """Place an order from the storefront. Prices are computed server-side."""
    data = request.get_json(force=True, silent=True) or {}
    nombre = (data.get('nombre') or '').strip()
    telefono = (data.get('telefono') or '').strip()
    items = data.get('items') or []

    if not nombre or not telefono:
        return jsonify({'ok': False, 'error': 'Nombre y teléfono son obligatorios'}), 400
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'El carrito está vacío'}), 400

    db = get_db()
    autocommit = db.autocommit
    db.autocommit = False
    try:
        factor = 1 + _get_margen(db) / 100.0
        final_items = []
        errors = []
        for it in items:
            sku = str(it.get('sku') or '').strip()
            try:
                qty = min(99, max(1, int(it.get('qty') or 1)))
            except (TypeError, ValueError):
                qty = 1
            if not sku:
                continue
            row = _query(db, '''SELECT p.sku, p.nombre, p.linea, p.stock,
                                       (ROUND((pr.precio * %s + 10) / 1000) * 1000 - 10) AS precio_venta
                                FROM products p
                                JOIN prices pr ON p.sku = pr.sku
                                     AND pr.import_date = (SELECT MAX(import_date) FROM imports)
                                WHERE p.sku = %s AND p.is_active = 1 AND p.stock > 0''',
                         [factor, sku]).fetchone()
            if not row:
                errors.append(f'{sku}: producto no disponible')
                continue
            if int(row['stock']) < qty:
                errors.append(f"{row['nombre']}: quedan {row['stock']} unidades")
                continue
            final_items.append({'sku': row['sku'], 'nombre': row['nombre'],
                                'linea': row['linea'], 'precio': int(row['precio_venta']),
                                'qty': qty})

        if errors:
            db.rollback()
            return jsonify({'ok': False, 'error': ' · '.join(errors)}), 400
        if not final_items:
            db.rollback()
            return jsonify({'ok': False, 'error': 'El carrito está vacío'}), 400

        total = sum(i['precio'] * i['qty'] for i in final_items)
        cur = db.cursor()
        cur.execute('''INSERT INTO orders (link_token, nombre, telefono, items_json, total, status)
                       VALUES (%s, %s, %s, %s, %s, %s)''',
                    ['WEB', nombre, telefono,
                     json.dumps(final_items, ensure_ascii=False), total, 'nuevo'])
        order_id = cur.lastrowid
        for i in final_items:
            cur.execute('UPDATE products SET stock = stock - %s WHERE sku = %s AND stock >= %s',
                        [i['qty'], i['sku'], i['qty']])
            if cur.rowcount == 0:
                raise Exception(f"Stock insuficiente para {i['nombre']}")
        cur.close()
        db.commit()
        return jsonify({'ok': True, 'order_id': order_id, 'total': total})
    except Exception as e:
        db.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.autocommit = autocommit
        db.close()


# ==================== ADMIN: inventario y configuración ====================

@app.route('/admin/inventario')
@require_login
def admin_inventario():
    query = request.args.get('q', '').strip()
    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = 50

    db = get_db()
    margen = _get_margen(db)
    factor = 1 + margen / 100.0

    sql = '''SELECT p.sku, p.nombre, p.linea, p.genero, p.stock, p.is_active, p.aroma_family,
                    pr.precio AS precio_mayor,
                    (ROUND((pr.precio * %s + 10) / 1000) * 1000 - 10) AS precio_venta
             FROM products p
             LEFT JOIN prices pr ON pr.sku = p.sku
                  AND pr.import_date = (SELECT MAX(import_date) FROM imports)'''
    params = [factor]
    if query:
        sql += ' WHERE (p.nombre LIKE %s OR p.sku LIKE %s OR p.linea LIKE %s)'
        like = f'%{query}%'
        params.extend([like, like, like])
    sql += ' ORDER BY p.linea, p.nombre'

    total = _query(db, f'SELECT COUNT(*) AS cnt FROM ({sql}) AS t', params).fetchone()['cnt']
    offset = (page - 1) * per_page
    products = _query(db, f'{sql} LIMIT %s OFFSET %s', params + [per_page, offset]).fetchall()
    total_pages = (total + per_page - 1) // per_page if total else 1
    for p in products:
        p['genero_norm'] = _norm_genero(p.get('genero'))

    activos = _query(db, 'SELECT COUNT(*) AS cnt FROM products WHERE is_active = 1 AND stock > 0').fetchone()['cnt']
    db.close()

    return render_template('inventario.html', products=products, query=query,
                           total=total, page=page, per_page=per_page, total_pages=total_pages,
                           margen_tienda=margen, activos=activos,
                           familias=sorted(AROMA_FAMILIES.keys()))


@app.route('/api/update_inventory', methods=['POST'])
@require_login
def update_inventory():
    data = request.get_json()
    sku = data.get('sku')
    stock = data.get('stock')
    is_active = data.get('is_active')
    aroma_family = data.get('aroma_family')

    db = get_db()
    try:
        _query(db, 'UPDATE products SET stock=%s, is_active=%s, aroma_family=%s WHERE sku=%s',
               [stock, 1 if is_active else 0, aroma_family, sku])
        db.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.rollback()
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        db.close()


@app.route('/admin/config', methods=['POST'])
@require_login
def admin_config():
    margen = request.form.get('margen_tienda', '').strip().replace(',', '.')
    try:
        m = float(margen)
        if not (0 <= m <= 500):
            raise ValueError()
    except ValueError:
        flash('El margen debe ser un número entre 0 y 500', 'error')
        return redirect(url_for('admin_inventario'))
    db = get_db()
    set_setting(db, 'margen_pct', int(m) if m == int(m) else m)
    db.close()
    flash(f'Margen de la tienda actualizado a {int(m) if m == int(m) else m}%', 'success')
    return redirect(url_for('admin_inventario'))


if __name__ == '__main__':
    init_db()
    app.run(debug=True)
else:
    init_db()

